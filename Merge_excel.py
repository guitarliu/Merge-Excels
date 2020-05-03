# -*- coding: utf-8 -*-
# @Time    : 2020/4/26 9:54
# @Author  : 刘勇
# @FileName: Merge_excel.py
# @Software: PyCharm
# @github ：https://github.com/guitarliu/

import os, re
import numpy
from openpyxl import *
from tkinter import *
from tkinter import filedialog, messagebox

def change_list(args):
    '''
    change the none-number ones of args's elements to 0
    '''
    for i in args:
        if not str(i).isdigit():
            args[args.index(i)] = 0
    return args

def merge_excels():
    '''
    extract the same contents and write them into a middle xlsx file
    extract the different contents of the second xlsx file and write them into a middle xlsx file
    '''
    if e_start.get() == "" or e_end.get() == "" or e_add_end.get() == "" or e_add_end.get() == "":
        messagebox.showwarning(title = "有空值", message = "输入为空，请重新输入！")
    else:
        wb1 = load_workbook(var_file_1)
        wb2 = load_workbook(var_file_2)
        wb_final = Workbook()
        ws1 = wb1.active
        ws2 = wb2.active
        ws_final = wb_final.active

        for row1 in ws1.rows:
            for row2 in ws2.rows:
                if [row1_content.value for row1_content in row1][(int(e_start.get())-1):int(e_end.get())] == \
                        [row2_content.value for row2_content in row2][(int(e_start.get())-1):int(e_end.get())]:
                    content1 = numpy.array(change_list([row1_content.value for row1_content in row1][(int(e_add_start.get())-1):int(e_add_end.get())]))
                    content2 = numpy.array(change_list([row2_content.value for row2_content in row2][(int(e_add_start.get())-1):int(e_add_end.get())]))
                    contents = list(content1 + content2)
                    final_contents = [row1_content.value for row1_content in row1][:(int(e_add_start.get())-1)] + contents + \
                                     [row1_content.value for row1_content in row1][int(e_add_end.get())::]
                    ws_final.append(final_contents)
                    wb_final.save(os.getcwd() + "/" + "middle.xlsx")
                    ws2.delete_rows(int(re.findall(r"\d+", str(row2[0]))[-1]))
                    wb2.save(var_file_2.split(".xlsx")[0] + "_middle.xlsx")

        # extract the different contents of the first xlsx file and write them into a middle xlsx file
        wb_middle = load_workbook(os.getcwd() + "/" + "middle.xlsx")
        wb1_a = load_workbook(var_file_1)
        ws_middle = wb_middle.active
        ws1_a = wb1_a.active

        for row1 in ws_middle.rows:
            for row2 in ws1_a.rows:
                if [row1_content.value for row1_content in row1][(int(e_start.get())-1):int(e_end.get())] == \
                        [row2_content.value for row2_content in row2][(int(e_start.get())-1):int(e_end.get())]:
                    ws1_a.delete_rows(int(re.findall(r"\d+", str(row2[0]))[-1]))
                    wb1_a.save(var_file_1.split(".xlsx")[0] + "_middle.xlsx")

        # write two middle xlsx files into the middle file
        wb_final = load_workbook(os.getcwd() + "/" + "middle.xlsx")
        wb1_middle = load_workbook(var_file_1.split(".xlsx")[0] + "_middle.xlsx")
        wb2_middle =load_workbook(var_file_2.split(".xlsx")[0] + "_middle.xlsx")
        ws_final = wb_final.active
        ws1_middle = wb1_middle.active
        ws2_middle = wb2_middle.active
        for row1 in ws1_middle.rows:
            ws_final.append([row1_content.value for row1_content in row1])
            wb_final.save(os.getcwd() + "/" + "final.xlsx")
        for row2 in ws2_middle.rows:
            ws_final.append([row2_content.value for row2_content in row2])
            wb_final.save(os.getcwd() + "/" + "final.xlsx")
        messagebox.showinfo(title="合并完成", message="%s和%s合并完成，请到当前文件夹下查看！（我，秦始皇，打钱！）" % (var_file_1, var_file_2))

def get_excel_filenames():
    '''get the filenames of two excel files'''
    global var_file_1
    global var_file_2
    filename= filedialog.askopenfilenames()
    if len(filename) != 2:
        messagebox.showwarning(title="选择失败",message="请重新选择两个Excel文件")
    var_filename.set(filename[0].split("/")[-1] + "\n" + filename[1].split("/")[-1])
    var_file_1 = filename[0]
    var_file_2 = filename[1]


def clear():
    '''clear the contents of all components'''
    var_filename.set("Excel文件\nExcel文件2")
    global var_file_1
    global var_file_2
    var_file_1 = ""
    var_file_2 = ""
    e_start.delete(0, END)
    e_end.delete(0, END)
    e_add_start.delete(0, END)
    e_add_end.delete(0, END)


# create the UI
root = Tk()
root.title('Excel合并')
var_filename = StringVar()
var_filename.set("Excel文件\nExcel文件2")
Button(root, text = "选择文件", command = get_excel_filenames).grid(row = 0, column = 0, padx = 1, pady = 1)
Label(root, text = "选择的文件为:", padx = 1, pady = 1).grid(row = 0, column = 1, padx = 1, pady = 1)
Label(root, textvariable = var_filename, padx = 1, pady = 1).grid(row = 0, column = 2, padx = 1, pady = 1)
Label(root, text = "比较项目开始列:", padx = 1, pady = 1).grid(row = 1, column = 0, padx = 1, pady = 1)
e_start = Entry(root)
e_start.grid(row = 1, column = 1, padx = 1, pady = 1)
Label(root, text = "比较项目结束列:", padx = 1, pady = 1).grid(row = 2, column = 0, padx = 1, pady = 1)
e_end = Entry(root)
e_end.grid(row = 2, column = 1, padx = 1, pady = 1)
Label(root, text = "加和项目开始列:", padx = 1, pady = 1).grid(row = 3, column = 0, padx = 1, pady = 1)
e_add_start = Entry(root)
e_add_start.grid(row = 3, column = 1, padx = 1, pady = 1)
Label(root, text = "加和项目结束列:", padx = 1, pady = 1).grid(row = 4, column = 0, padx = 1, pady = 1)
e_add_end = Entry(root)
e_add_end.grid(row = 4, column = 1, padx = 1, pady = 1)
Label(root, text = "说明：只能比较同一行连续列的内容", padx = 1, pady = 1).grid(row=5, column=1)
Label(root, text = "例如比较第1/2行，4/5列的内容", padx = 1, pady = 1).grid(row=6, column=1)
Label(root, text = "对对应行第7/../13列内容相加").grid(row=7, column=1, pady = 3)
Button(root, text = "开始合并", command = merge_excels).grid(row=9, column = 0, padx = 1, pady = 5)
Button(root, text = "清   空", command = clear).grid(row=9, column = 1, padx = 1, pady = 5)
Button(root, text = "退   出", command = quit).grid(row=9, column = 2, padx = 1, pady = 5)
mainloop()
